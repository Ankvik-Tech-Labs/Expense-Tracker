"""
Excel exporter for portfolio data.

This module generates Excel reports with multiple sheets containing portfolio data.
"""
from datetime import datetime
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def create_portfolio_excel(
    holdings_df: pd.DataFrame,
    snapshots_df: pd.DataFrame,
    summary: dict,
    export_date: Optional[datetime] = None
) -> BytesIO:
    """
    Create an Excel file with portfolio data.

    :param holdings_df: DataFrame containing all holdings
    :type holdings_df: pd.DataFrame
    :param snapshots_df: DataFrame containing monthly snapshots
    :type snapshots_df: pd.DataFrame
    :param summary: Dictionary with portfolio summary statistics
    :type summary: dict
    :param export_date: Date for the export (defaults to now)
    :type export_date: Optional[datetime]
    :returns: BytesIO buffer containing the Excel file
    :rtype: BytesIO
    """
    if export_date is None:
        export_date = datetime.now()

    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    currency_format = '₹#,##0.00'
    percent_format = '0.00%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _create_summary_sheet(ws_summary, summary, export_date, header_font, header_fill)

    # Sheet 2: Stocks
    ws_stocks = wb.create_sheet("Stocks")
    stocks_df = holdings_df[holdings_df['type'] == 'stock'].copy() if not holdings_df.empty else pd.DataFrame()
    _create_holdings_sheet(ws_stocks, stocks_df, "Stocks Holdings", header_font, header_fill, thin_border)

    # Sheet 3: Mutual Funds
    ws_mf = wb.create_sheet("Mutual Funds")
    mf_df = holdings_df[holdings_df['type'] == 'mutual_fund'].copy() if not holdings_df.empty else pd.DataFrame()
    _create_holdings_sheet(ws_mf, mf_df, "Mutual Funds Holdings", header_font, header_fill, thin_border)

    # Sheet 4: Monthly Trend
    ws_trend = wb.create_sheet("Monthly Trend")
    _create_trend_sheet(ws_trend, snapshots_df, header_font, header_fill, thin_border)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _create_summary_sheet(ws, summary, export_date, header_font, header_fill):
    """Create the summary sheet with portfolio overview."""
    # Title
    ws['A1'] = "Portfolio Summary Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:C1')

    ws['A2'] = f"Generated on: {export_date.strftime('%d %b %Y %H:%M')}"
    ws['A2'].font = Font(italic=True, color='666666')

    # Summary data
    data = [
        ('Total Portfolio Value', summary.get('total_value', 0)),
        ('Total Invested', summary.get('total_invested', 0)),
        ('Total P&L', summary.get('total_pl', 0)),
        ('Total P&L %', summary.get('total_pl_pct', 0) / 100),
        ('', ''),
        ('Stocks Value', summary.get('stocks_value', 0)),
        ('Mutual Funds Value', summary.get('mf_value', 0)),
        ('US Stocks Value', summary.get('us_stocks_value', 0)),
    ]

    row = 4
    for label, value in data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(bold=True)

        if label == 'Total P&L %':
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = '0.00%'
        elif label:
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = '₹#,##0.00'

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _create_holdings_sheet(ws, holdings_df, title, header_font, header_fill, border):
    """Create a holdings sheet (Stocks or Mutual Funds)."""
    # Title
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)

    if holdings_df.empty:
        ws['A3'] = "No data available"
        return

    # Prepare data
    display_df = holdings_df[[
        'name', 'symbol', 'units', 'avg_price', 'invested_value',
        'current_price', 'current_value', 'unrealized_pl', 'unrealized_pl_pct'
    ]].copy()

    display_df.columns = [
        'Name', 'Symbol', 'Units', 'Avg Price', 'Invested Value',
        'Current Price', 'Current Value', 'P&L', 'P&L %'
    ]

    # Write headers
    for col_idx, col_name in enumerate(display_df.columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Write data
    for row_idx, row_data in enumerate(display_df.values, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            # Format currency columns
            if col_idx in [4, 5, 6, 7, 8]:
                cell.number_format = '₹#,##0.00'
            elif col_idx == 9:
                cell.number_format = '0.00%'
                cell.value = value / 100 if value else 0

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 15


def _create_trend_sheet(ws, snapshots_df, header_font, header_fill, border):
    """Create the monthly trend sheet."""
    # Title
    ws['A1'] = "Monthly Portfolio Trend"
    ws['A1'].font = Font(bold=True, size=14)

    if snapshots_df.empty:
        ws['A3'] = "No historical data available"
        return

    # Prepare data
    df = snapshots_df.sort_values('snapshot_date', ascending=False).copy()

    display_df = df[[
        'snapshot_date', 'total_value', 'total_invested', 'total_pl', 'total_pl_pct',
        'stocks_value', 'mf_value'
    ]].copy()

    display_df.columns = [
        'Date', 'Total Value', 'Invested', 'P&L', 'P&L %',
        'Stocks', 'Mutual Funds'
    ]

    # Format date
    display_df['Date'] = display_df['Date'].dt.strftime('%b %Y')

    # Write headers
    for col_idx, col_name in enumerate(display_df.columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Write data
    for row_idx, row_data in enumerate(display_df.values, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            # Format currency columns
            if col_idx in [2, 3, 4, 6, 7]:
                cell.number_format = '₹#,##0.00'
            elif col_idx == 5:
                cell.number_format = '0.00%'
                cell.value = value / 100 if value else 0

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15
